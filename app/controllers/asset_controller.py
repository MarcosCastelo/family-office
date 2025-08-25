from flask import jsonify, request, abort
from app.models.asset import Asset
from app.schema.asset_schema import AssetSchema
from app.config.extensions import db
from app.decorators.family_access import require_family
from app.services.asset_validation_service import AssetValidationService
from sqlalchemy import cast, String

asset_schema = AssetSchema()
assets_schema = AssetSchema(many=True)

def list_assets_controller(req):
    family_id = req.args.get("family_id")
    if not family_id:
        return jsonify({"error": "family_id é obrigatório"}), 400
    
    try:
        family_id = int(family_id)
    except (ValueError, TypeError):
        return jsonify({"error": "family_id deve ser um número válido"}), 400
    
    # Verificar acesso à família
    from flask_jwt_extended import get_jwt_identity
    from app.models.user import User
    user_id = get_jwt_identity()
    user = db.session.get(User, user_id)
    if not user or not any(f.id == family_id for f in user.families):
        return jsonify({"error": "Acesso à familia negado"}), 403
    
    assets = Asset.query.filter_by(family_id=family_id).all()
    return jsonify(assets_schema.dump(assets)), 200

def get_asset_controller(asset_id):
    family_id = request.args.get("family_id")
    if not family_id:
        return jsonify({"error": "family_id é obrigatório"}), 400
    
    try:
        family_id = int(family_id)
    except (ValueError, TypeError):
        return jsonify({"error": "family_id deve ser um número válido"}), 400
    
    # Verificar acesso à família
    from flask_jwt_extended import get_jwt_identity
    from app.models.user import User
    user_id = get_jwt_identity()
    user = db.session.get(User, user_id)
    if not user or not any(f.id == family_id for f in user.families):
        return jsonify({"error": "Acesso à familia negado"}), 403
    
    asset = db.session.get(Asset, asset_id)
    if not asset:
        abort(404)
    # Verificar se o ativo pertence à família
    if asset.family_id != family_id:
        return jsonify({"error": "Acesso negado"}), 403
    
    return jsonify(asset_schema.dump(asset)), 200

def create_asset_controller(req):
    data = req.get_json()
    errors = asset_schema.validate(data)
    if errors:
        return jsonify(errors), 400
    family_id = data.get("family_id")
    if not family_id:
        return jsonify({"error": "family_id é obrigatório"}), 400
    # Verificar acesso à família
    from flask_jwt_extended import get_jwt_identity
    from app.models.user import User
    user_id = get_jwt_identity()
    user = db.session.get(User, user_id)
    if not user or not any(f.id == int(family_id) for f in user.families):
        return jsonify({"error": "Acesso à familia negado"}), 403
    # Validação do nome do ativo
    name_validation = AssetValidationService.validate_asset_name(data["name"])
    if not name_validation[0]:
        return jsonify({"error": name_validation[1]}), 400
    
    # Validação de duplicidade de nome/ticker na família
    query = Asset.query.filter_by(family_id=family_id, name=data["name"])
    if data.get("asset_type") == "renda_variavel" and data.get("details", {}).get("ticker"):
        query = query.filter(cast(Asset.details["ticker"], String) == str(data["details"]["ticker"]))
    if query.first():
        return jsonify({"error": "Já existe ativo com este nome/ticker nesta família"}), 400
    
    # Validação de campos obrigatórios por classe
    details = data.get("details", {})
    asset_type = data.get("asset_type")
    
    # Validação específica por tipo de ativo
    if asset_type == "renda_fixa":
        fixed_income_validation = AssetValidationService.validate_fixed_income_fields(details)
        if not fixed_income_validation[0]:
            return jsonify({"error": fixed_income_validation[1]}), 400
    
    elif asset_type == "renda_variavel":
        ticker = details.get("ticker", "")
        if not ticker:
            return jsonify({"error": "Renda variável exige ticker"}), 400
        
        # Normalizar e validar ticker
        normalized_ticker = AssetValidationService.normalize_ticker(ticker, asset_type, 'B3')
        ticker_validation = AssetValidationService.validate_ticker(normalized_ticker, asset_type, 'B3')
        if not ticker_validation[0]:
            return jsonify({"error": ticker_validation[1]}), 400
        
        # Atualizar ticker normalizado
        details["ticker"] = normalized_ticker
    
    elif asset_type == "criptomoeda":
        coin_id = details.get("coin_id", "")
        if not coin_id:
            return jsonify({"error": "Criptomoeda exige ID da moeda"}), 400
        
        # Normalizar e validar coin_id
        normalized_coin_id = AssetValidationService.normalize_ticker(coin_id, asset_type)
        coin_validation = AssetValidationService.validate_ticker(normalized_coin_id, asset_type)
        if not coin_validation[0]:
            return jsonify({"error": coin_validation[1]}), 400
        
        # Atualizar coin_id normalizado
        details["coin_id"] = normalized_coin_id
    
    elif asset_type == "moeda_estrangeira":
        currency = details.get("currency", "")
        if not currency:
            return jsonify({"error": "Moeda estrangeira exige código da moeda"}), 400
        
        currency_validation = AssetValidationService.validate_ticker(currency, asset_type)
        if not currency_validation[0]:
            return jsonify({"error": currency_validation[1]}), 400
    asset = Asset(**data)
    db.session.add(asset)
    db.session.commit()
    # Geração automática de alertas
    gerar_alertas_ativos(asset.family_id)
    return jsonify(asset_schema.dump(asset)), 201

def update_asset_controller(asset_id, req):
    family_id = req.args.get("family_id")
    if not family_id:
        return jsonify({"error": "family_id é obrigatório"}), 400
    
    try:
        family_id = int(family_id)
    except (ValueError, TypeError):
        return jsonify({"error": "family_id deve ser um número válido"}), 400
    
    # Verificar acesso à família
    from flask_jwt_extended import get_jwt_identity
    from app.models.user import User
    user_id = get_jwt_identity()
    user = db.session.get(User, user_id)
    if not user or not any(f.id == family_id for f in user.families):
        return jsonify({"error": "Acesso à familia negado"}), 403
    
    asset = db.session.get(Asset, asset_id)
    if not asset:
        abort(404)
    # Verificar se o ativo pertence à família
    if asset.family_id != family_id:
        return jsonify({"error": "Acesso negado"}), 403
    
    data = req.get_json()
    errors = asset_schema.validate(data, partial=True)
    if errors:
        return jsonify(errors), 400
    for key, value in data.items():
        setattr(asset, key, value)
    db.session.commit()
    return jsonify(asset_schema.dump(asset)), 200

def delete_asset_controller(asset_id):
    family_id = request.args.get("family_id")
    if not family_id:
        return jsonify({"error": "family_id é obrigatório"}), 400
    
    try:
        family_id = int(family_id)
    except (ValueError, TypeError):
        return jsonify({"error": "family_id deve ser um número válido"}), 400
    
    # Verificar acesso à família
    from flask_jwt_extended import get_jwt_identity
    from app.models.user import User
    user_id = get_jwt_identity()
    user = db.session.get(User, user_id)
    if not user or not any(f.id == family_id for f in user.families):
        return jsonify({"error": "Acesso à familia negado"}), 403
    
    asset = db.session.get(Asset, asset_id)
    if not asset:
        abort(404)
    # Verificar se o ativo pertence à família
    if asset.family_id != family_id:
        return jsonify({"error": "Acesso negado"}), 403
    
    db.session.delete(asset)
    db.session.commit()
    return jsonify({"msg": "Ativo removido com sucesso"}), 204

def upload_assets_controller(req):
    import csv
    import io
    from flask import jsonify
    from werkzeug.datastructures import FileStorage
    from marshmallow import ValidationError
    import openpyxl
    file: FileStorage = req.files.get('file')
    if not file or not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        return jsonify({'error': 'Arquivo CSV ou XLSX obrigatório'}), 400
    try:
        assets = []
        if file.filename.endswith('.csv'):
            content = file.read().decode('utf-8')
            reader = csv.DictReader(content.splitlines())
            for row in reader:
                import json
                if 'details' in row and isinstance(row['details'], str):
                    try:
                        row['details'] = json.loads(row['details'])
                    except Exception:
                        row['details'] = {}
                if 'family_id' in row:
                    try:
                        row['family_id'] = int(row['family_id'])
                    except Exception:
                        return jsonify({'error': 'family_id inválido'}), 400
                errors = asset_schema.validate(row)
                if errors:
                    return jsonify(errors), 400
                assets.append(row)
        elif file.filename.endswith('.xlsx'):
            file.seek(0)
            wb = openpyxl.load_workbook(filename=io.BytesIO(file.read()), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row_cells in ws.iter_rows(min_row=2):
                row = {headers[i]: (cell.value if cell.value is not None else "") for i, cell in enumerate(row_cells)}
                import json
                if 'details' in row and isinstance(row['details'], str):
                    try:
                        row['details'] = json.loads(row['details'])
                    except Exception:
                        row['details'] = {}
                if 'family_id' in row:
                    try:
                        row['family_id'] = int(row['family_id'])
                    except Exception:
                        return jsonify({'error': 'family_id inválido'}), 400
                errors = asset_schema.validate(row)
                if errors:
                    return jsonify(errors), 400
                assets.append(row)
        else:
            return jsonify({'error': 'Formato de arquivo não suportado'}), 400
        # Autorização: checar se usuário pertence à família
        from flask_jwt_extended import get_jwt_identity
        from app.models.user import User
        user_id = get_jwt_identity()
        user = db.session.get(User, user_id)
        for asset in assets:
            fam_id = asset['family_id']
            if not user or not any(f.id == fam_id for f in user.families):
                return jsonify({'error': 'Acesso à familia negado'}), 403
        # Persistir
        created = []
        for asset in assets:
            obj = Asset(**asset)
            db.session.add(obj)
            db.session.flush()
            created.append(asset_schema.dump(obj))
        db.session.commit()
        return jsonify(created), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 400

def upload_pdf_assets_controller(req):
    from flask import jsonify
    from werkzeug.datastructures import FileStorage
    from app.services.gemini_ocr_service import extract_assets_from_pdf
    
    # Get the requested family_id from query parameters
    family_id = req.args.get("family_id")
    if not family_id:
        return jsonify({'error': 'family_id é obrigatório'}), 400
    try:
        family_id = int(family_id)
    except (ValueError, TypeError):
        return jsonify({'error': 'family_id deve ser um número válido'}), 400
    
    # Check file
    file: FileStorage = req.files.get('file')
    if not file or not file.filename.endswith('.pdf'):
        return jsonify({'error': 'Arquivo PDF obrigatório'}), 400
    
    # Check user access to the family BEFORE calling Gemini API
    from flask_jwt_extended import get_jwt_identity
    from app.models.user import User
    user_id = get_jwt_identity()
    user = db.session.get(User, user_id)
    if not user or not any(f.id == family_id for f in user.families):
        return jsonify({'error': 'Acesso à familia negado'}), 403
    
    try:
        # Call Gemini API to extract assets
        assets = extract_assets_from_pdf(file)
        print(f"DEBUG: Gemini extracted assets: {assets}")  # Debug log
        
    except Exception as e:
        return jsonify({'error': f'Erro na API Gemini: {str(e)}'}), 502
    
    # Validate the returned format
    if not isinstance(assets, list):
        return jsonify({'error': 'Resposta da API Gemini não é uma lista de ativos'}), 400
    
    # Add default values for missing fields
    from datetime import date
    for asset in assets:
        if 'acquisition_date' not in asset or not asset['acquisition_date']:
            asset['acquisition_date'] = date.today().isoformat()
            print(f"DEBUG: Added default acquisition_date: {asset['acquisition_date']}")
        
        if 'details' not in asset or not asset['details']:
            asset['details'] = {}
            print(f"DEBUG: Added default details: {{}}")
    
    # Process and save assets
    created = []
    for i, asset in enumerate(assets):
        print(f"DEBUG: Processing asset {i}: {asset}")  # Debug log
        
        # Override family_id with the requested one to ensure consistency
        asset['family_id'] = family_id
        
        # Provide default values for missing fields
        if 'acquisition_date' not in asset or not asset['acquisition_date']:
            from datetime import date
            asset['acquisition_date'] = date.today().isoformat()
            print(f"DEBUG: Added default acquisition_date: {asset['acquisition_date']}")
        
        if 'details' not in asset or not asset['details']:
            asset['details'] = {}
            print(f"DEBUG: Added default details: {{}}")
        
        # Validate asset data
        errors = asset_schema.validate(asset)
        if errors:
            print(f"DEBUG: Validation errors for asset {i}: {errors}")  # Debug log
            return jsonify(errors), 400
        
        # Create and save asset
        obj = Asset(**asset)
        db.session.add(obj)
        db.session.flush()
        created.append(asset_schema.dump(obj))
        print(f"DEBUG: Successfully created asset {i}")  # Debug log
    
    db.session.commit()
    return jsonify(created), 201

def get_asset_risk_controller(asset_id):
    from flask import jsonify, request
    family_id = request.args.get("family_id")
    if not family_id:
        return jsonify({"error": "family_id é obrigatório"}), 400
    try:
        family_id = int(family_id)
    except (ValueError, TypeError):
        return jsonify({"error": "family_id deve ser um número válido"}), 400
    # Verificar acesso à família
    from flask_jwt_extended import get_jwt_identity
    from app.models.user import User
    user_id = get_jwt_identity()
    user = db.session.get(User, user_id)
    if not user or not any(f.id == family_id for f in user.families):
        return jsonify({"error": "Acesso à familia negado"}), 403
    asset = db.session.get(Asset, asset_id)
    if not asset:
        return jsonify({"error": "Ativo não encontrado"}), 404
    if asset.family_id != family_id:
        return jsonify({"error": "Acesso negado"}), 403
    # Cálculo real de risco
    details = asset.details or {}
    risco_mercado = "baixo"
    risco_liquidez = "baixo"
    risco_concentracao = "baixo"
    risco_credito = "baixo"
    risco_cambial = "baixo"
    risco_juridico_fiscal = "baixo"
    score_governanca = details.get("governanca", 100)
    classificacao_final = "baixo"
    # Renda variável, cripto, valor alto, governança baixa = alto risco
    if asset.asset_type == "renda_variavel":
        risco_mercado = "alto"
        if details.get("setor", "").lower() == "cripto":
            risco_mercado = "crítico"
        if asset.value > 500000:
            risco_concentracao = "alto"
        if score_governanca < 30:
            score_governanca = int(score_governanca)
            classificacao_final = "crítico"
        elif score_governanca < 60:
            classificacao_final = "alto"
        else:
            classificacao_final = "alto"
    # Liquidez
    if details.get("liquidez", "alta").lower() == "baixa":
        risco_liquidez = "alto"
        classificacao_final = "alto"
    # Renda fixa conservadora
    if asset.asset_type == "renda_fixa" and details.get("indexador", "").lower() in ["ipca", "selic"]:
        risco_mercado = "baixo"
        classificacao_final = "baixo"
    # Ajuste final: se algum risco for crítico, classifica como crítico
    if risco_mercado == "crítico" or risco_liquidez == "crítico":
        classificacao_final = "crítico"
    risco = {
        "id": asset.id,
        "name": asset.name,
        "risco_mercado": risco_mercado,
        "risco_liquidez": risco_liquidez,
        "risco_concentracao": risco_concentracao,
        "risco_credito": risco_credito,
        "risco_cambial": risco_cambial,
        "risco_juridico_fiscal": risco_juridico_fiscal,
        "score_governanca": score_governanca,
        "classificacao_final": classificacao_final
    }
    return jsonify(risco), 200

def gerar_alertas_ativos(family_id):
    from app.models.asset import Asset
    from app.models.alert import Alert
    from app.config.extensions import db
    # Remove alertas antigos de concentração e liquidez (um a um)
    for alerta in db.session.query(Alert).filter_by(family_id=family_id, tipo="concentracao").all():
        db.session.delete(alerta)
    for alerta in db.session.query(Alert).filter_by(family_id=family_id, tipo="liquidez").all():
        db.session.delete(alerta)
    db.session.commit()
    # Concentração: alerta se algum ativo > 30% do total
    ativos = Asset.query.filter_by(family_id=family_id).all()
    print(f"[DEBUG ALERTAS] Ativos na família {family_id}: {[{'id': a.id, 'name': a.name, 'value': a.value} for a in ativos]}")
    total = sum(a.value for a in ativos)
    for ativo in ativos:
        if len(ativos) > 1 and total > 0 and ativo.value / total > 0.3:
            alerta = Alert(
                family_id=family_id,
                asset_id=ativo.id,
                tipo="concentracao",
                mensagem=f"Ativo '{ativo.name}' representa mais de 30% da carteira.",
                severidade="warning"
            )
            db.session.add(alerta)
    def is_iliquido(a):
        d = a.details or {}
        return str(d.get("liquidez", "alta")).lower() == "baixa"
    iliquidos = [a for a in ativos if is_iliquido(a)]
    total_iliquido = sum(a.value for a in iliquidos)
    if total > 0 and total_iliquido / total > 0.5:
        alerta = Alert(
            family_id=family_id,
            tipo="liquidez",
            mensagem="Mais de 50% da carteira está em ativos ilíquidos.",
            severidade="danger"
        )
        db.session.add(alerta)
    db.session.commit()